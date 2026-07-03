#!/usr/bin/env python3
"""Génère un classeur Excel récapitulatif à partir d'une commande (1..N plats).

Usage :
  python generer.py "curry-pois-chiches:20, veloute-potimarron:20" [--nom NOM] [--sortie FICHIER.xlsx]
  python generer.py --commande commandes/semaine-2026-07-06.md [--nom semaine] [--sortie ...]

Une "commande" = des couples plat:couverts. En mode --commande, le script récupère tous les
motifs `slug: nombre` du fichier dont le slug correspond à une recette existante (les doublons
d'un même plat sont cumulés).

4 onglets : Récap commande | Liste de courses | Apports par plat | Allergènes par plat.
Calcul déterministe à partir de recettes/*.md et data/*.csv. Valeurs INDICATIVES (cf. README).
"""
import argparse
import csv
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

# 14 allergènes INCO (ordre officiel) + libellés courts pour les colonnes Excel
INCO = ["gluten", "crustaces", "oeufs", "poissons", "arachides", "soja", "lait",
        "fruits-a-coque", "celeri", "moutarde", "sesame", "sulfites", "lupin", "mollusques"]

RAYON_ORDER = ["Fruits & légumes", "Boucherie / Volaille", "Poissonnerie", "Crémerie / Frais",
               "Épicerie salée", "Épicerie sucrée", "Surgelés", "Boissons", "Divers / à vérifier"]

NUTRIENTS = ["kcal", "proteines_g", "glucides_g", "sucres_g", "lipides_g",
             "satures_g", "fibres_g"]

VOLUME_UNITS = {"ml", "verre"}


def trouver_racine(depart: Path) -> Path:
    """Remonte jusqu'au dossier contenant recettes/ et data/."""
    for p in [depart] + list(depart.parents):
        if (p / "recettes").is_dir() and (p / "data").is_dir():
            return p
    raise SystemExit("Racine du harness introuvable (dossiers recettes/ et data/ requis).")


def lire_csv_commentaire(chemin: Path):
    """Lit un CSV en ignorant les lignes de commentaire (#)."""
    with open(chemin, encoding="utf-8") as f:
        lignes = [l for l in f if not l.lstrip().startswith("#")]
    return list(csv.DictReader(lignes))


def charger_donnees(racine: Path):
    nutrition = {}
    for row in lire_csv_commentaire(racine / "data" / "nutrition-100g.csv"):
        ing = row["ingredient"]
        nutrition[ing] = {k: float(row[k]) for k in NUTRIENTS}

    allergenes = {}
    for row in lire_csv_commentaire(racine / "data" / "allergenes.csv"):
        pres = set() if row["allergenes"].strip() in ("", "aucun") else \
            {x.strip() for x in row["allergenes"].split(";") if x.strip()}
        trac = set() if (row.get("traces_possibles") or "").strip() in ("", "aucun") else \
            {x.strip() for x in row["traces_possibles"].split(";") if x.strip()}
        allergenes[row["ingredient"]] = (pres, trac)

    unite_g, piece_g = {}, {}
    for row in lire_csv_commentaire(racine / "data" / "conversions.csv"):
        if row["type"] == "unite":
            unite_g[row["cle"]] = float(row["grammes"])
        elif row["type"] == "piece":
            piece_g[row["cle"]] = float(row["grammes"])

    rayons = {r["ingredient"]: r["rayon"]
              for r in lire_csv_commentaire(racine / "data" / "rayons.csv")}
    return nutrition, allergenes, unite_g, piece_g, rayons


def parse_recette(chemin: Path):
    texte = chemin.read_text(encoding="utf-8")
    m = re.search(r"^---\n(.*?)\n---\n(.*)$", texte, re.S)
    if not m:
        raise ValueError(f"Frontmatter manquant : {chemin}")
    front_brut, corps = m.group(1), m.group(2)
    front = {}
    for ligne in front_brut.splitlines():
        if ":" in ligne:
            k, v = ligne.split(":", 1)
            front[k.strip()] = v.strip()

    ingredients = []
    dans_section = False
    for ligne in corps.splitlines():
        if ligne.startswith("## Ingrédients"):
            dans_section = True
            continue
        if ligne.startswith("## ") and dans_section:
            break
        if dans_section and ligne.strip().startswith("- "):
            ingredients.append(ligne.strip()[2:].strip())
    return front, ingredients


def grammes(ligne: str, unite_g, piece_g):
    """'250 g pois chiches' -> (grammes, nom_ingredient, est_volume)."""
    m = re.match(r"^([0-9]+(?:[.,][0-9]+)?)\s+(.*)$", ligne)
    if not m:
        return None
    qte = float(m.group(1).replace(",", "."))
    reste = m.group(2).strip()
    for u in sorted(list(unite_g) + ["pièce"], key=len, reverse=True):
        if reste.startswith(u + " "):
            ing = reste[len(u) + 1:].strip()
            if u == "pièce":
                return qte * piece_g.get(ing, 100.0), ing, False
            return qte * unite_g[u], ing, (u in VOLUME_UNITS)
    # pas d'unité reconnue -> on considère une pièce
    return qte * piece_g.get(reste, 100.0), reste, False


def parse_commande(args, racine):
    """Retourne [(slug, couverts), ...] cumulés."""
    cumul = defaultdict(float)
    recettes_dispo = {p.stem for p in (racine / "recettes").glob("*.md")
                      if p.stem not in ("README", "_format")}
    if args.commande:
        texte = Path(args.commande).read_text(encoding="utf-8")
        for slug, n in re.findall(r"([a-z0-9][a-z0-9-]+)\s*:\s*(\d+)", texte):
            if slug in recettes_dispo:
                cumul[slug] += int(n)
    if args.spec:
        for bloc in args.spec.split(","):
            bloc = bloc.strip()
            if not bloc:
                continue
            slug, _, n = bloc.partition(":")
            cumul[slug.strip()] += int(n.strip())
    if not cumul:
        raise SystemExit("Commande vide : fournir 'slug:couverts,...' ou --commande FICHIER.")
    return sorted(cumul.items())


def fmt_quantite(g, est_volume, ing, piece_g):
    if est_volume:
        base = f"{g/1000:.2f} L" if g >= 1000 else f"{g:.0f} ml"
    else:
        base = f"{g/1000:.2f} kg" if g >= 1000 else f"{g:.0f} g"
    if ing in piece_g and piece_g[ing] > 0:
        base += f" (≈ {round(g/piece_g[ing])} pièces)"
    return base


def slugify(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


def main():
    ap = argparse.ArgumentParser(description="Génère le classeur Excel d'une commande traiteur.")
    ap.add_argument("spec", nargs="?", help="'slug:couverts, slug:couverts, ...'")
    ap.add_argument("--commande", help="Fichier (plan/semaine) contenant des 'slug: couverts'.")
    ap.add_argument("--nom", help="Nom de la commande (titre + nom de fichier).")
    ap.add_argument("--sortie", help="Chemin du .xlsx de sortie.")
    ap.add_argument("--racine", help="Racine du harness (auto-détectée par défaut).")
    args = ap.parse_args()

    racine = Path(args.racine).resolve() if args.racine else trouver_racine(Path(__file__).resolve())
    nutrition, allergenes, unite_g, piece_g, rayons = charger_donnees(racine)
    commande = parse_commande(args, racine)

    nom = args.nom or (Path(args.commande).stem if args.commande else f"commande-{date.today()}")
    sortie = Path(args.sortie) if args.sortie else (racine / "commandes" / f"{slugify(nom)}.xlsx")
    sortie.parent.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise SystemExit("openpyxl requis : pip install openpyxl")

    # --- agrégations ---
    recap = []                       # (nom, type, couverts, portions, facteur, cout)
    courses = defaultdict(float)     # ingredient -> grammes totaux
    courses_vol = {}                 # ingredient -> est_volume
    apports = []                     # (nom, {nutriment: par_portion}, couverts)
    allerg_plat = []                 # (nom, present:set, traces:set)
    inconnus = set()

    for slug, couverts in commande:
        chemin = racine / "recettes" / f"{slug}.md"
        if not chemin.exists():
            inconnus.add(slug)
            continue
        front, ings = parse_recette(chemin)
        portions = float(front.get("portions", 1) or 1)
        facteur = couverts / portions
        cout = float(front.get("cout-estime-eur", 0) or 0) * facteur
        recap.append((front.get("nom", slug), front.get("type-repas", ""),
                      couverts, portions, facteur, cout))

        par_portion = {k: 0.0 for k in NUTRIENTS}
        present, traces = set(), set()
        for li in ings:
            res = grammes(li, unite_g, piece_g)
            if not res:
                continue
            g, ing, est_vol = res
            courses[ing] += g * facteur
            courses_vol[ing] = courses_vol.get(ing, False) or est_vol
            if ing in nutrition:
                for k in NUTRIENTS:
                    par_portion[k] += g / 100.0 * nutrition[ing][k]
            else:
                inconnus.add(ing)
            p, t = allergenes.get(ing, (set(), set()))
            present |= p
            traces |= t
        par_portion = {k: v / portions for k, v in par_portion.items()}
        apports.append((front.get("nom", slug), par_portion, couverts))
        allerg_plat.append((front.get("nom", slug), present, traces - present))

    wb = Workbook()
    gras = Font(bold=True)
    entete_fill = PatternFill("solid", fgColor="DDEBF7")
    centre = Alignment(horizontal="center")

    def style_entete(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = gras
            cell.fill = entete_fill
        ws.freeze_panes = "A2"

    # Onglet 1 : Récap commande
    ws = wb.active
    ws.title = "Récap commande"
    ws.append(["Plat", "Type", "Couverts", "Portions réf.", "Facteur", "Coût (€)"])
    total_couverts = total_cout = 0
    for n, typ, couv, port, fac, cout in recap:
        ws.append([n, typ, couv, port, round(fac, 2), round(cout, 2)])
        total_couverts += couv
        total_cout += cout
    ws.append(["TOTAL", "", total_couverts, "", "", round(total_cout, 2)])
    for c in range(1, 7):
        ws.cell(row=ws.max_row, column=c).font = gras
    style_entete(ws, 6)

    # Onglet 2 : Liste de courses
    ws = wb.create_sheet("Liste de courses")
    ws.append(["Rayon", "Ingrédient", "Quantité (achat)", "Grammes/ml totaux"])
    for ing in sorted(courses, key=lambda i: (RAYON_ORDER.index(rayons.get(i, "Divers / à vérifier"))
                                              if rayons.get(i, "Divers / à vérifier") in RAYON_ORDER
                                              else 99, i)):
        g = courses[ing]
        ws.append([rayons.get(ing, "Divers / à vérifier"), ing,
                   fmt_quantite(g, courses_vol.get(ing, False), ing, piece_g), round(g)])
    style_entete(ws, 4)

    # Onglet 3 : Apports par plat (par portion) + totaux commande
    ws = wb.create_sheet("Apports par plat")
    cols = ["Plat", "kcal/portion", "Protéines (g)", "Glucides (g)", "dont sucres (g)",
            "Lipides (g)", "dont saturés (g)", "Fibres (g)"]
    ws.append(cols)
    totaux = {k: 0.0 for k in NUTRIENTS}
    for n, pp, couv in apports:
        ws.append([n, round(pp["kcal"]), round(pp["proteines_g"], 1), round(pp["glucides_g"], 1),
                   round(pp["sucres_g"], 1), round(pp["lipides_g"], 1), round(pp["satures_g"], 1),
                   round(pp["fibres_g"], 1)])
        for k in NUTRIENTS:
            totaux[k] += pp[k] * couv
    ws.append(["TOTAL COMMANDE (servi)", round(totaux["kcal"]), round(totaux["proteines_g"]),
               round(totaux["glucides_g"]), round(totaux["sucres_g"]), round(totaux["lipides_g"]),
               round(totaux["satures_g"]), round(totaux["fibres_g"])])
    for c in range(1, len(cols) + 1):
        ws.cell(row=ws.max_row, column=c).font = gras
    style_entete(ws, len(cols))

    # Onglet 4 : Allergènes par plat
    ws = wb.create_sheet("Allergènes par plat")
    ws.append(["Plat"] + INCO)
    for n, present, traces in allerg_plat:
        ligne = [n]
        for a in INCO:
            ligne.append("X" if a in present else ("trace" if a in traces else "—"))
        ws.append(ligne)
    for r in range(2, ws.max_row + 1):
        for c in range(2, len(INCO) + 2):
            ws.cell(row=r, column=c).alignment = centre
    style_entete(ws, len(INCO) + 1)

    # largeurs de colonnes lisibles
    for ws in wb.worksheets:
        for col in ws.columns:
            largeur = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(largeur + 2, 8), 40)

    wb.save(sortie)

    print(f"✅ Classeur écrit : {sortie}")
    print(f"   Plats : {len(recap)} | Couverts : {total_couverts} | Coût estimé : {round(total_cout, 2)} €")
    print(f"   Onglets : {[ws.title for ws in wb.worksheets]}")
    if inconnus:
        print(f"   ⚠️ Ingrédients/plats hors référentiel (estimés ou ignorés) : {sorted(inconnus)}")


if __name__ == "__main__":
    sys.exit(main())
